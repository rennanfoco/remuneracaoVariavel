"""
Gera arquivos de exemplo na pasta samples/ com nomes e formatos que o loader aceita.
Cobre todos os grupos de cargo, incluindo proporcionalidade e inelegibilidade.

Uso:
  python generate_samples.py
"""

import csv
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("Instale dependências: pip install -r requirements.txt")
    raise

SAMPLES_DIR = Path("samples")
SAMPLES_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Unidades e regionais
# Usa códigos de loja reais (presentes em rv/loader.py::_NONREV_PARA_TOTVS) para
# que o arquivo de NONREV (que usa o código sem sufixo, ex: "AJU") seja traduzido
# corretamente para o código TOTVS (ex: "AJU10") usado nos demais arquivos.
# ---------------------------------------------------------------------------
UNIT_CODES = ["AJU", "BPS", "CGR", "CNF", "CWB", "FLN", "FOR", "GIG", "GYN", "JPA"]
UNIDADES = [
    ("AJU10", "REGIONAL SUL"),
    ("BPS10", "REGIONAL SUL"),
    ("CGR10", "REGIONAL SUL"),
    ("CNF10", "REGIONAL SUL"),
    ("CWB10", "REGIONAL SUL"),
    ("FLN10", "REGIONAL NORTE"),
    ("FOR10", "REGIONAL NORTE"),
    ("GIG10", "REGIONAL NORTE"),
    ("GYN10", "REGIONAL NORTE"),
    ("JPA10", "REGIONAL NORTE"),
]

# NPS por unidade (valor direto 0–100)
NPS_UNIDADE = {
    "AJU10": 80.5, "BPS10": 78.2, "CGR10": 75.0, "CNF10": 70.0, "CWB10": 79.5,
    "FLN10": 77.3, "FOR10": 78.8, "GIG10": 74.0, "GYN10": 76.5, "JPA10": 82.0,
}

# NONREV por unidade (% — ex: 2.30 = 2,30%)
NONREV_UNIDADE = {
    "AJU10": 2.30, "BPS10": 2.80, "CGR10": 4.00, "CNF10": 2.50, "CWB10": 3.20,
    "FLN10": 2.40, "FOR10": 2.70, "GIG10": 3.80, "GYN10": 2.50, "JPA10": 2.10,
}

# % Manutenção Preventiva por unidade
PREVENTIVA_UNIDADE = {
    "AJU10": 99.5, "BPS10": 97.8, "CGR10": 95.0, "CNF10": 99.2, "CWB10": 98.1,
    "FLN10": 99.8, "FOR10": 96.5, "GIG10": 97.3, "GYN10": 99.0, "JPA10": 100.0,
}

# % Bate Pátio por unidade (1.0 = 100%)
BATE_PATIO_UNIDADE = {
    "AJU10": 1.00, "BPS10": 0.95, "CGR10": 1.00, "CNF10": 0.88, "CWB10": 1.00,
    "FLN10": 1.00, "FOR10": 0.90, "GIG10": 1.00, "GYN10": 1.00, "JPA10": 0.85,
}

# DMA-alvo por índice de atendente no ciclo (em R$/diária)
# Ciclo repetido; gera diferentes faixas (faixa2≥40, faixa1=[38,40), zero<38)
DMA_ALVO = [41.0, 39.0, 36.0, 42.5, 38.5]

# Percentual de atingimento de faturamento por índice (1.00 = meta exata)
FAT_ATING = [1.07, 1.02, 0.95, 1.10, 1.03]


# ---------------------------------------------------------------------------
# Colaboradores (colaboradores_rm.csv — nome fixo)
# ---------------------------------------------------------------------------

def _colaboradores():
    rows = []
    mat = 1000

    def add(cargo, unidade, regional, salario, dias=30, n=1):
        nonlocal mat
        for _ in range(n):
            rows.append({
                "matricula":        str(mat).zfill(6),
                "nome":             f"Colaborador {mat}",
                "cargo":            cargo,
                "unidade":          unidade,
                "regional":         regional,
                "salario_base":     str(salario).replace(".", ","),
                "dias_trabalhados": str(dias),
            })
            mat += 1

    for und, reg in UNIDADES:
        add("ATENDENTE",         und, reg, 1800.00, n=5)
        add("ATENDENTE LIDER",   und, reg, 2500.00)
        add("GESTOR",            und, reg, 4000.00)
        add("LIDER DE FROTA",    und, reg, 3000.00)
        add("LIDER DE PATIO",    und, reg, 3000.00)
        add("OPERADOR DE FROTAS", und, reg, 2200.00, n=2)
        add("AGENTE RECEPTIVO",  und, reg, 1700.00)
        add("ASSISTENTE ADM",    und, reg, 1900.00)
        add("AUX. DE ATENDIMENTO", und, reg, 1500.00)

    # Casos especiais — proporcionalidade (admissão no mês)
    und0, reg0 = UNIDADES[0]
    add("ATENDENTE", und0, reg0, 1800.00, dias=15)
    add("ATENDENTE", und0, reg0, 1800.00, dias=10)

    for reg in ("REGIONAL SUL", "REGIONAL NORTE"):
        und = next(u for u, r in UNIDADES if r == reg)
        rows.append({
            "matricula":        str(mat).zfill(6),
            "nome":             f"Regional {reg}",
            "cargo":            "REGIONAL",
            "unidade":          und,
            "regional":         reg,
            "salario_base":     "8000,00",
            "dias_trabalhados": "30",
        })
        mat += 1

    return rows


# ---------------------------------------------------------------------------
# Coral — dados brutos de locações (dma-accumulated*.csv)
# DMA e faturamento são calculados pelo loader a partir das linhas individuais.
# Colunas obrigatórias: LOGIN, LOJA, VALOR VENDIDO, DIAS, PREMIAVEL
# PREMIAVEL: 'yes' | 'no' | 'voucher-required'  (só 'no' é excluído do DMA)
# No arquivo real, cada contrato tem uma linha por item opcional e a coluna
# DIAS só vem preenchida na linha do item base (LDW) — o loader soma DIAS
# por login para obter o denominador do DMA.
# ---------------------------------------------------------------------------

def _coral(colaboradores):
    """
    Retorna (linhas_coral, {login: mat, ...}, {login: meta_faturamento}).
    Gera 10 transações por atendente/lider: 8 com PREMIAVEL='yes', 2 com 'no'.
    """
    atendentes = [
        c for c in colaboradores
        if c["cargo"] in ("ATENDENTE", "ATENDENTE LIDER")
    ]

    rows = []
    login_to_mat = {}
    meta_por_login = {}
    idx = 0

    for c in atendentes:
        mat = c["matricula"]
        login = f"at{mat}"
        und = c["unidade"]
        login_to_mat[login] = mat

        dma_alvo = DMA_ALVO[idx % len(DMA_ALVO)]
        fat_ating = FAT_ATING[idx % len(FAT_ATING)]

        # 8 transações premiáveis + 2 não premiáveis
        for _ in range(8):
            rows.append({
                "LOGIN":         login,
                "LOJA":          und,
                "VALOR VENDIDO": f"{dma_alvo:.2f}".replace(".", ","),
                "DIAS":          "1",
                "PREMIAVEL":     "yes",
            })
        for _ in range(2):
            rows.append({
                "LOGIN":         login,
                "LOJA":          und,
                "VALOR VENDIDO": f"{dma_alvo:.2f}".replace(".", ","),
                "DIAS":          "1",
                "PREMIAVEL":     "no",
            })

        # meta de faturamento: fat_realizado = 10 * dma_alvo; meta = fat / ating
        fat_realizado = 10 * dma_alvo
        meta = fat_realizado / fat_ating
        meta_por_login[login] = round(meta, 2)
        idx += 1

    return rows, login_to_mat, meta_por_login


# ---------------------------------------------------------------------------
# NPS — já vem agregado por loja na origem (data*.xlsx)
# Formato real: linhas de Ano/Mês antes do header, depois Loja | NPS | %.
# ---------------------------------------------------------------------------

def _nps():
    return {und: NPS_UNIDADE[und] for und, _ in UNIDADES}


# ---------------------------------------------------------------------------
# Preventiva — % Conformidade por loja, já calculado na origem
# Aba 'Análise de Preventivas', bloco 'ANÁLISE DETALHADA POR LOJA':
# header Loja | ... | % Conformidade | ..., dados por loja, linha TOTAL final.
# ---------------------------------------------------------------------------

def _preventiva():
    return {und: PREVENTIVA_UNIDADE[und] for und, _ in UNIDADES}


# ---------------------------------------------------------------------------
# Bate Pátio / Fechamento — por loja (Fechamento*.xlsx)
# Colunas: FILIAL, % POR LOJA  (valores em fração: 1.0 = 100%)
# ---------------------------------------------------------------------------

def _bate_patio():
    return [
        {"FILIAL": und, "% POR LOJA": str(BATE_PATIO_UNIDADE[und]).replace(".", ",")}
        for und, _ in UNIDADES
    ]


# ---------------------------------------------------------------------------
# Meta Faturamento — por login (Importação Planilha*.csv)
# Estrutura: header PT → linha EN (skipped pelo loader) → dados
# Separador: ; | Encoding: utf-8-sig
# ---------------------------------------------------------------------------

def _meta_faturamento(meta_por_login):
    rows = []
    # Primeira linha de dados = nomes de campo em inglês (o loader vai pular)
    rows.append({"Usuário": "user_login", "Meta de Faturamento": "billing_goal"})
    for login, meta in meta_por_login.items():
        rows.append({
            "Usuário":             login,
            "Meta de Faturamento": f"{meta:.2f}".replace(".", ","),
        })
    return rows


# ---------------------------------------------------------------------------
# Mapeamento Login Coral → Matrícula RM (Lista de Logins*.xlsx)
# Colunas: Login, Matrícula
# ---------------------------------------------------------------------------

def _mapeamento(login_to_mat):
    return [
        {"Login": login, "Matrícula": mat}
        for login, mat in login_to_mat.items()
    ]


# ---------------------------------------------------------------------------
# NONREV — por loja (nonrev*.xlsx)
# Arquivo tem linhas de filtro antes do header real; loader busca linha com "Loja".
# Valores em decimal (0.023 = 2.3%); loja SEM sufixo "10" (loader acrescenta).
# ---------------------------------------------------------------------------

def _nonrev():
    # Códigos-base (sem sufixo), igual ao formato real da área de Frotas —
    # o loader traduz para o código TOTVS via _NONREV_PARA_TOTVS.
    codes = list(UNIT_CODES)
    vals  = [NONREV_UNIDADE[u + "10"] / 100.0 for u in UNIT_CODES]
    return codes, vals


# ---------------------------------------------------------------------------
# Gravação dos arquivos
# ---------------------------------------------------------------------------

def _csv(rows, filename, fieldnames=None, encoding="utf-8-sig"):
    path = SAMPLES_DIR / filename
    if not rows:
        print(f"  [aviso] {filename} — nenhuma linha gerada.")
        return
    if fieldnames is None:
        fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding=encoding) as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
    print(f"  {filename} ({len(rows)} linhas)")


def _xlsx(rows, filename):
    path = SAMPLES_DIR / filename
    wb = Workbook()
    ws = wb.active
    if not rows:
        print(f"  [aviso] {filename} — nenhuma linha gerada.")
        wb.save(path)
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    wb.save(path)
    print(f"  {filename} ({len(rows)} linhas)")


def _xlsx_nonrev(unit_codes, valores, filename):
    """
    Estrutura especial do NONREV:
      linha 0 → filtro (texto livre, primeira coluna)
      linha 1 → vazia
      linha 2 → "Loja" | "NONREV %"   (header real — o loader detecta aqui)
      linha 3+ → dados
    """
    path = SAMPLES_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.append(["Filtro: Todas as Lojas", ""])
    ws.append(["", ""])
    ws.append(["Loja", "NONREV %"])
    for code, val in zip(unit_codes, valores):
        ws.append([code, val])
    wb.save(path)
    print(f"  {filename} ({len(unit_codes)} lojas)")


def _xlsx_nps(unit_pct: dict, filename):
    """
    Estrutura do NPS (data*.xlsx): linhas de Ano/Mês antes do header real,
    depois 'Loja' | 'NPS' | '%', uma linha por loja, e um rodapé de filtros.
    """
    path = SAMPLES_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.append(["Ano", 2026, 2026])
    ws.append(["Mês", "jun", "jun"])
    ws.append(["Loja", "NPS", "%"])
    for und, nps in unit_pct.items():
        ws.append([und, nps, 0.30])
    ws.append([])
    ws.append(["Filtros aplicados:\nAno é 2026"])
    wb.save(path)
    print(f"  {filename} ({len(unit_pct)} lojas)")


def _xlsx_preventiva(unit_pct: dict, filename):
    """
    Estrutura da aba 'Análise de Preventivas' (bloco 'ANÁLISE DETALHADA POR LOJA'):
      linhas de título antes do header real
      header: 'Loja' | ... | '% Conformidade' | ...
      uma linha por loja
      linha 'TOTAL' encerrando o bloco
    """
    path = SAMPLES_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "Análise de Preventivas"
    ws.append(["ANÁLISE DE PREVENTIVAS — BASE GERENCIAL (BONIFICAÇÃO)"])
    ws.append([])
    ws.append(["Loja", "Regional", "Total Placas", "Ok", "Pendentes", "% Conformidade", "Meta"])
    for und, pct in unit_pct.items():
        frac = pct / 100.0
        ws.append([und, "", 20, round(frac * 20), 20 - round(frac * 20), frac, 0.9901])
    ws.append(["TOTAL", "", "", "", "", "", ""])
    wb.save(path)
    print(f"  {filename} ({len(unit_pct)} lojas)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Gerando arquivos de exemplo em samples/...\n")

    colaboradores = _colaboradores()
    coral_rows, login_to_mat, meta_por_login = _coral(colaboradores)
    nonrev_codes, nonrev_vals = _nonrev()

    # colaboradores_rm.csv — nome FIXO (não usa glob)
    _csv(colaboradores, "colaboradores_rm.csv")

    # dma-accumulated_sample.csv — glob: dma-accumulated*.csv
    _csv(coral_rows, "dma-accumulated_sample.csv",
         fieldnames=["LOGIN", "LOJA", "VALOR VENDIDO", "DIAS", "PREMIAVEL"])

    # data_nps.xlsx — glob: data*.xlsx
    _xlsx_nps(_nps(), "data_nps.xlsx")

    # Preventivas_amostra.xlsx — glob: Preventivas*.xlsx
    _xlsx_preventiva(_preventiva(), "Preventivas_amostra.xlsx")

    # Fechamento_amostra.xlsx — glob: Fechamento*.xlsx
    _xlsx(_bate_patio(), "Fechamento_amostra.xlsx")

    # Importacao Planilha_amostra.csv — glob: Importa??o Planilha*.csv
    # (o ? no glob corresponde a 'ç')
    _csv(_meta_faturamento(meta_por_login), "Importação Planilha_amostra.csv",
         fieldnames=["Usuário", "Meta de Faturamento"])

    # Lista de Logins_amostra.xlsx — glob: Lista de Logins*.xlsx
    _xlsx(_mapeamento(login_to_mat), "Lista de Logins_amostra.xlsx")

    # nonrev_amostra.xlsx — glob: nonrev*.xlsx
    _xlsx_nonrev(nonrev_codes, nonrev_vals, "nonrev_amostra.xlsx")

    n_colab = len(colaboradores)
    n_atend = sum(1 for c in colaboradores if c["cargo"] in ("ATENDENTE", "ATENDENTE LIDER"))
    print(f"\nTotal de colaboradores gerados   : {n_colab}")
    print(f"Atendentes com dados Coral       : {n_atend}")
    print(f"Linhas brutas no arquivo Coral   : {len(coral_rows)}")
    print("\nPara executar o calculo com os dados de exemplo:")
    print("  python main.py --competencia 2026-06 --sample")


if __name__ == "__main__":
    main()
