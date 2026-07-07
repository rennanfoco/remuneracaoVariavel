"""
Cria o arquivo referencias.xlsx com os parâmetros iniciais do processo de RV.
Execute uma vez para criar o arquivo. Depois edite diretamente no Excel.

Uso: python criar_referencias.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT = Path("referencias.xlsx")


def _estilizar_cabecalho(ws, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 20


def _escrever(ws, cabecalho, linhas):
    ws.append(cabecalho)
    _estilizar_cabecalho(ws, len(cabecalho))
    for linha in linhas:
        ws.append(list(linha))


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------
    # Regras_Calculo
    # Relaciona cada grupo de cargo com seus indicadores e faixas.
    # Formato LONGO: uma linha por FAIXA, não uma coluna por faixa — para
    # adicionar uma faixa3 (ou mais), basta adicionar mais linhas, sem
    # alterar nenhum código.
    #
    # Colunas:
    #   grupo        — nome interno do grupo (atendente, lider_frota, etc.), sempre minúsculo
    #   indicador    — nome da métrica (dma, nps, faturamento, preventiva, nonrev, bate_patio)
    #   faixa        — número da faixa (1, 2, 3, ...). Número maior = melhor faixa.
    #   valor_min    — valor mínimo da faixa (deixar em branco se definido em Metas_Mensais,
    #                  ou se essa faixa for a mais alta em direcao=maior)
    #   valor_max    — valor máximo (exclusivo) da faixa (deixar em branco se essa faixa for
    #                  a mais alta em direcao=maior, ou a mais baixa em direcao=menor)
    #   pct          — % do salário/teto creditado nessa faixa (ex: 0,15 = 15%)
    #   direcao      — "maior" (maior é melhor) ou "menor" (menor é melhor, ex: NONREV) —
    #                  repetir em todas as linhas do mesmo grupo+indicador
    #   chave_meta   — chave em Metas_Mensais para limites variáveis por mês; deixar em
    #                  branco para limites fixos — repetir em todas as linhas
    #
    # Obs sobre faturamento: thresholds em % de atingimento (100 = meta exata, 105 = 5% acima da meta).
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Regras_Calculo")
    _escrever(ws,
        ["grupo", "indicador", "faixa", "valor_min", "valor_max", "pct", "direcao", "chave_meta"],
        [
            # ── Modelo % do Salário Base ──────────────────────────────
            # Atendente: DMA e Faturamento individuais; NPS da unidade
            ("atendente",       "dma",         1, 38,   40,   0.15, "maior", None),
            ("atendente",       "dma",         2, 40,   None, 0.45, "maior", None),
            ("atendente",       "nps",         1, 74,   77,   0.05, "maior", None),
            ("atendente",       "nps",         2, 77,   None, 0.10, "maior", None),
            ("atendente",       "faturamento", 1, 100,  105,  0.15, "maior", None),
            ("atendente",       "faturamento", 2, 105,  None, 0.45, "maior", None),
            # Atendente Lider: todos os indicadores da unidade
            ("atendente_lider", "dma",         1, 38,   40,   0.15, "maior", None),
            ("atendente_lider", "dma",         2, 40,   None, 0.45, "maior", None),
            ("atendente_lider", "nps",         1, 74,   77,   0.05, "maior", None),
            ("atendente_lider", "nps",         2, 77,   None, 0.10, "maior", None),
            ("atendente_lider", "faturamento", 1, 100,  105,  0.15, "maior", None),
            ("atendente_lider", "faturamento", 2, 105,  None, 0.45, "maior", None),
            # Gestor: todos os indicadores da unidade
            ("gestor",          "dma",         1, 38,   40,   0.15, "maior", None),
            ("gestor",          "dma",         2, 40,   None, 0.45, "maior", None),
            ("gestor",          "nps",         1, 74,   77,   0.05, "maior", None),
            ("gestor",          "nps",         2, 77,   None, 0.10, "maior", None),
            ("gestor",          "faturamento", 1, 100,  105,  0.15, "maior", None),
            ("gestor",          "faturamento", 2, 105,  None, 0.45, "maior", None),
            # Regional: consolidado das unidades da regional
            ("regional",        "dma",         1, 38,   40,   0.15, "maior", None),
            ("regional",        "dma",         2, 40,   None, 0.45, "maior", None),
            ("regional",        "nps",         1, 74,   77,   0.05, "maior", None),
            ("regional",        "nps",         2, 77,   None, 0.10, "maior", None),
            ("regional",        "faturamento", 1, 100,  105,  0.15, "maior", None),
            ("regional",        "faturamento", 2, 105,  None, 0.45, "maior", None),

            # ── Modelo Teto Fixo ──────────────────────────────────────
            # Lider Frota: NPS (meta mensal), Preventiva (fixo), NONREV (meta mensal)
            ("lider_frota",     "nps",         1, None, None, 0.20, "maior", "nps_lider_frota"),
            ("lider_frota",     "nps",         2, None, None, 0.40, "maior", "nps_lider_frota"),
            ("lider_frota",     "preventiva",  1, 97,   99,   0.15, "maior", None),
            ("lider_frota",     "preventiva",  2, 99,   None, 0.30, "maior", None),
            ("lider_frota",     "nonrev",      1, None, None, 0.15, "menor", "nonrev"),
            ("lider_frota",     "nonrev",      2, None, None, 0.30, "menor", "nonrev"),
            # Lider Patio: NPS (meta mensal), NONREV (meta mensal), Bate Patio (fixo)
            ("lider_patio",     "nps",         1, None, None, 0.20, "maior", "nps_lider_patio"),
            ("lider_patio",     "nps",         2, None, None, 0.40, "maior", "nps_lider_patio"),
            ("lider_patio",     "nonrev",      1, None, None, 0.15, "menor", "nonrev"),
            ("lider_patio",     "nonrev",      2, None, None, 0.30, "menor", "nonrev"),
            ("lider_patio",     "bate_patio",  2, 100,  None, 0.30, "maior", None),
            # Operador Frota: NPS (meta mensal), Preventiva (fixo), NONREV (meta mensal)
            ("operador_frota",  "nps",         1, None, None, 0.20, "maior", "nps_operador_frota"),
            ("operador_frota",  "nps",         2, None, None, 0.40, "maior", "nps_operador_frota"),
            ("operador_frota",  "preventiva",  1, 97,   99,   0.15, "maior", None),
            ("operador_frota",  "preventiva",  2, 99,   None, 0.30, "maior", None),
            ("operador_frota",  "nonrev",      1, None, None, 0.15, "menor", "nonrev"),
            ("operador_frota",  "nonrev",      2, None, None, 0.30, "menor", "nonrev"),
            # Outros Cargos: apenas NPS (meta mensal)
            ("outros",          "nps",         1, None, None, 0.50, "maior", "nps_outros"),
            ("outros",          "nps",         2, None, None, 1.00, "maior", "nps_outros"),
            # Mecanico: Preventiva (fixo) e NONREV (meta mensal) — modelo percentual (salário)
            ("mecanico",        "preventiva",  1, 96,   99,   0.15, "maior", None),
            ("mecanico",        "preventiva",  2, 99,   None, 0.25, "maior", None),
            ("mecanico",        "nonrev",      1, None, None, 0.15, "menor", "nonrev"),
            ("mecanico",        "nonrev",      2, None, None, 0.25, "menor", "nonrev"),
        ]
    )

    # ------------------------------------------------------------------
    # Grupos
    # Define, para cada grupo usado em Regras_Calculo, se a base do prêmio é
    # o salário do colaborador (percentual) ou um teto fixo em R$ (teto_fixo,
    # ver aba Tetos). Todo grupo que aparece em Regras_Calculo PRECISA de uma
    # linha aqui — o cálculo é cancelado no início se faltar algum.
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Grupos")
    _escrever(ws,
        ["grupo", "modelo"],
        [
            ("atendente",       "percentual"),
            ("atendente_lider", "percentual"),
            ("gestor",          "percentual"),
            ("regional",        "percentual"),
            ("mecanico",        "percentual"),
            ("lider_frota",     "teto_fixo"),
            ("lider_patio",     "teto_fixo"),
            ("operador_frota",  "teto_fixo"),
            ("outros",          "teto_fixo"),
        ]
    )

    # ------------------------------------------------------------------
    # Cargos
    # Mapeamento cargo (exato como vem da API.04 / colaboradores_rm.csv) → grupo_calculo.
    # Execute: python test_totvs_api.py --mes MM --ano AAAA --list-cargos
    # para ver todos os cargos distintos e confirmar os nomes exatos.
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Cargos")
    _escrever(ws,
        ["cargo", "grupo_calculo"],
        [
            # ── Atendimento ───────────────────────────────────────────
            ("ATENDENTE",                              "atendente"),
            ("ATENDENTE I",                            "atendente"),
            ("ATENDENTE PL",                           "atendente"),
            ("ATENDENTE SR",                           "atendente"),
            ("ATENDENTE SR III",                       "atendente"),
            ("AGENTE RECEPTIVO",                       "atendente"),
            ("AGENTE DE LOCACAO",                      "atendente"),
            ("OPERADOR DE ATENDIMENTO I",              "atendente"),
            ("OPERADOR DE ATENDIMENTO AO CLIENTE",     "atendente"),
            ("AUXILIAR DE ATENDIMENTO",                "atendente"),
            # ── Líderes de loja ───────────────────────────────────────
            ("ATENDENTE LIDER",                        "atendente_lider"),
            ("ATENDENTE LIDER II",                     "atendente_lider"),
            ("SUPERVISOR DE LOJA",                     "atendente_lider"),
            ("SUPERVISOR DE LOJA I",                   "atendente_lider"),
            ("SUPERVISOR DE LOJA II",                  "atendente_lider"),
            # ── Gestores de loja ──────────────────────────────────────
            ("GESTOR",                                 "gestor"),
            ("GERENTE DE LOJA I",                      "gestor"),
            ("GERENTE DE LOJA II",                     "gestor"),
            # ── Regionais ─────────────────────────────────────────────
            ("REGIONAL",                               "regional"),
            ("GERENTE REGIONAL I",                     "regional"),
            ("GERENTE REGIONAL II",                    "regional"),
            ("GERENTE REGIONAL III",                   "regional"),
            # ── Frota – líderes ───────────────────────────────────────
            ("LIDER DE FROTA",                         "lider_frota"),
            ("LIDER DE FROTAS",                        "lider_frota"),
            ("LIDER DE FROTAS II",                     "lider_frota"),
            ("COORDENADOR DE FROTAS I",                "lider_frota"),
            # ── Frota – pátio ─────────────────────────────────────────
            ("LIDER DE PATIO",                         "lider_patio"),
            ("LIDER DE PATIO II",                      "lider_patio"),
            # ── Frota – operadores ────────────────────────────────────
            ("OPERADOR DE FROTAS",                     "operador_frota"),
            ("OPERADOR DE FROTAS I",                   "operador_frota"),
            ("OPERADOR DE FROTAS SR",                  "operador_frota"),
            # ── Outros ────────────────────────────────────────────────
            ("ASSISTENTE ADM",                         "outros"),
            ("ASSISTENTE ADM FINANCEIRO",              "outros"),
            ("ANALISTA ADMINISTRATIVO",                "outros"),
            ("ANALISTA ADM",                           "outros"),
            ("AUX ADM",                                "outros"),
            ("AUX. ADM",                               "outros"),
            ("AUX DE ATENDIMENTO",                     "outros"),
            ("AUX. DE ATENDIMENTO",                    "outros"),
            ("AUXILIAR DE SERVICOS GERAIS",            "outros"),
            ("AUXILIAR DE SERVICOS GERAIS PART-TIME",  "outros"),
            ("AUX DE SERVICOS GERAIS",                 "outros"),
            ("MOTORISTA D",                            "outros"),
            ("MOTORISTA CATEGORIA D",                  "outros"),
            ("MECANICO JR",                            "outros"),
            ("MECANICO PL",                            "outros"),
            ("MECANICO SR",                            "outros"),
            ("LAVADOR",                                "outros"),
            ("EXECUTIVO DE CONTAS",                    "outros"),
            ("JOVEM APRENDIZ",                         "outros"),
            ("ASSISTENTE DE PREVENCAO DE PERDAS",      "outros"),
            ("ANALISTA DE PREVENCAO DE PERDAS",        "outros"),
            ("HEAD DE PREVENCAO DE PERDAS E FACILITIES", "outros"),
            ("ASSISTENTE DE MONITORAMENTO",            "outros"),
            ("COORDENADOR TECNICO PLENO",              "outros"),
            ("COORDENADOR DE FACILITIES",              "outros"),
            ("HEAD DE CORPORATE FINANCE",              "outros"),
            ("GERENTE DE SINISTROS JUNIOR",            "outros"),
        ]
    )

    # ------------------------------------------------------------------
    # Tetos
    # Valor máximo de prêmio em R$ por cargo (modelo teto fixo).
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Tetos")
    _escrever(ws,
        ["cargo", "teto"],
        [
            # ── lider_frota ───────────────────────────────────────────
            ("LIDER DE FROTA",                         700.00),
            ("LIDER DE FROTAS",                        700.00),
            ("LIDER DE FROTAS II",                     700.00),
            ("COORDENADOR DE FROTAS I",                700.00),
            # ── lider_patio ───────────────────────────────────────────
            ("LIDER DE PATIO",                         700.00),
            ("LIDER DE PATIO II",                      700.00),
            # ── operador_frota ────────────────────────────────────────
            ("OPERADOR DE FROTAS",                     500.00),
            ("OPERADOR DE FROTAS I",                   500.00),
            ("OPERADOR DE FROTAS SR",                  500.00),
            # ── outros ────────────────────────────────────────────────
            ("AGENTE RECEPTIVO",                       200.00),
            ("AGENTE DE LOCACAO",                      200.00),
            ("OPERADOR DE ATENDIMENTO I",              200.00),
            ("OPERADOR DE ATENDIMENTO AO CLIENTE",     200.00),
            ("AUXILIAR DE ATENDIMENTO",                200.00),
            ("ASSISTENTE ADM",                         200.00),
            ("ASSISTENTE ADM FINANCEIRO",              200.00),
            ("ANALISTA ADMINISTRATIVO",                200.00),
            ("ANALISTA ADM",                           200.00),
            ("AUX ADM",                                200.00),
            ("AUX. ADM",                               200.00),
            ("AUX DE ATENDIMENTO",                     200.00),
            ("AUX. DE ATENDIMENTO",                    200.00),
            ("MOTORISTA D",                            200.00),
            ("MOTORISTA CATEGORIA D",                  200.00),
            ("EXECUTIVO DE CONTAS",                    200.00),
            ("MECANICO JR",                            200.00),
            ("MECANICO PL",                            200.00),
            ("MECANICO SR",                            200.00),
            ("LAVADOR",                                200.00),
            ("JOVEM APRENDIZ",                         200.00),
            ("ASSISTENTE DE PREVENCAO DE PERDAS",      200.00),
            ("ANALISTA DE PREVENCAO DE PERDAS",        200.00),
            ("HEAD DE PREVENCAO DE PERDAS E FACILITIES", 200.00),
            ("ASSISTENTE DE MONITORAMENTO",            200.00),
            ("COORDENADOR TECNICO PLENO",              200.00),
            ("COORDENADOR DE FACILITIES",              200.00),
            ("HEAD DE CORPORATE FINANCE",              200.00),
            ("GERENTE DE SINISTROS JUNIOR",            200.00),
            ("AUXILIAR DE SERVICOS GERAIS",            100.00),
            ("AUXILIAR DE SERVICOS GERAIS PART-TIME",  100.00),
            ("AUX DE SERVICOS GERAIS",                 100.00),
            ("AUX. DE SERVICOS GERAIS",                100.00),
        ]
    )

    # ------------------------------------------------------------------
    # Metas_Mensais
    # NPS e NONREV variam a cada mês. Formato LONGO: uma linha por faixa —
    # mesma lógica de Regras_Calculo. Adicione um bloco de linhas por mês/chave;
    # para uma faixa3, basta adicionar mais uma linha por chave.
    # NPS    → faixa mais alta sem valor_max (maior é melhor)
    # NONREV → faixa mais alta sem valor_min (menor é melhor)
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Metas_Mensais")
    _escrever(ws,
        ["mes", "chave", "faixa", "valor_min", "valor_max"],
        [
            # fev/26 ──────────────────────────────────────────────────
            ("2026-02", "nps_lider_frota",    1, 75.0, 77.0),
            ("2026-02", "nps_lider_frota",    2, 77.0, None),
            ("2026-02", "nps_lider_patio",    1, 75.0, 77.0),
            ("2026-02", "nps_lider_patio",    2, 77.0, None),
            ("2026-02", "nps_operador_frota", 1, 75.0, 77.0),
            ("2026-02", "nps_operador_frota", 2, 77.0, None),
            ("2026-02", "nps_outros",         1, 75.0, 77.0),
            ("2026-02", "nps_outros",         2, 77.0, None),
            ("2026-02", "nonrev",             1, 2.61, 3.50),
            ("2026-02", "nonrev",             2, None, 2.60),
            # mar/26 ──────────────────────────────────────────────────
            ("2026-03", "nps_lider_frota",    1, 75.0, 77.0),
            ("2026-03", "nps_lider_frota",    2, 77.0, None),
            ("2026-03", "nps_lider_patio",    1, 76.0, 78.0),
            ("2026-03", "nps_lider_patio",    2, 78.0, None),
            ("2026-03", "nps_operador_frota", 1, 76.0, 78.0),
            ("2026-03", "nps_operador_frota", 2, 78.0, None),
            ("2026-03", "nps_outros",         1, 76.0, 78.0),
            ("2026-03", "nps_outros",         2, 78.0, None),
            ("2026-03", "nonrev",             1, 2.61, 3.50),
            ("2026-03", "nonrev",             2, None, 2.60),
            # abr/26 ──────────────────────────────────────────────────
            ("2026-04", "nps_lider_frota",    1, 76.0, 78.0),
            ("2026-04", "nps_lider_frota",    2, 78.0, None),
            ("2026-04", "nps_lider_patio",    1, 76.0, 78.0),
            ("2026-04", "nps_lider_patio",    2, 78.0, None),
            ("2026-04", "nps_operador_frota", 1, 76.0, 78.0),
            ("2026-04", "nps_operador_frota", 2, 78.0, None),
            ("2026-04", "nps_outros",         1, 76.0, 78.0),
            ("2026-04", "nps_outros",         2, 78.0, None),
            ("2026-04", "nonrev",             1, 2.51, 3.50),
            ("2026-04", "nonrev",             2, None, 2.50),
            # mai/26 ──────────────────────────────────────────────────
            ("2026-05", "nps_lider_frota",    1, 77.0, 79.0),
            ("2026-05", "nps_lider_frota",    2, 79.0, None),
            ("2026-05", "nps_lider_patio",    1, 77.0, 79.0),
            ("2026-05", "nps_lider_patio",    2, 79.0, None),
            ("2026-05", "nps_operador_frota", 1, 77.0, 79.0),
            ("2026-05", "nps_operador_frota", 2, 79.0, None),
            ("2026-05", "nps_outros",         1, 77.0, 79.0),
            ("2026-05", "nps_outros",         2, 79.0, None),
            ("2026-05", "nonrev",             1, 2.51, 3.50),
            ("2026-05", "nonrev",             2, None, 2.50),
            # jun/26 ──────────────────────────────────────────────────
            ("2026-06", "nps_lider_frota",    1, 78.0, 80.0),
            ("2026-06", "nps_lider_frota",    2, 80.0, None),
            ("2026-06", "nps_lider_patio",    1, 77.0, 79.0),
            ("2026-06", "nps_lider_patio",    2, 79.0, None),
            ("2026-06", "nps_operador_frota", 1, 77.0, 79.0),
            ("2026-06", "nps_operador_frota", 2, 79.0, None),
            ("2026-06", "nps_outros",         1, 77.0, 79.0),
            ("2026-06", "nps_outros",         2, 79.0, None),
            ("2026-06", "nonrev",             1, 2.51, 3.50),
            ("2026-06", "nonrev",             2, None, 2.50),
        ]
    )

    # ------------------------------------------------------------------
    # TOTVS — parâmetros de integração
    # ------------------------------------------------------------------
    ws = wb.create_sheet("TOTVS")
    _escrever(ws,
        ["parametro", "valor", "descricao"],
        [
            ("cod_evento_rv", "1001", "Codigo do evento de RV na folha TOTVS"),
            ("cod_coligada",  "1",    "Codigo da coligada no TOTVS"),
            ("separador",     ";",    "Separador de campos no arquivo TXT"),
        ]
    )

    wb.save(OUTPUT)
    print(f"Arquivo criado: {OUTPUT.resolve()}")
    print("Abra no Excel para ajustar os valores antes de rodar main.py.")


if __name__ == "__main__":
    main()
